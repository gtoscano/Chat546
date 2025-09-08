# python manage.py runscript load_members

#from researcher.models import Researcher
#from student.models import Student
import random
import openpyxl
import ast
import sys
import codecs
import fileinput
from django.conf import settings
from django.contrib.auth import get_user_model
from core.models import *
from experts.models import *


def run():
    from email.utils import parseaddr
    ## initializing the xlsx
    filename = 'scripts/profiles.xlsx'

    xlsx = openpyxl.load_workbook(filename)
    sheet_names = xlsx.get_sheet_names()
    errors_list = []
    usr_dict = {}
    experts = Expert.objects.all()

    users = get_user_model().objects.all()
    emails = []
    for usr in users:
        emails.append(usr.email)
    User = get_user_model()

    institutions = Institution.objects.all()
    current_institutions = []
    for institution in institutions:
        current_institutions.append(institution.acronym)

    if 'Institutions' in sheet_names:
        tab = xlsx['Institutions']
        row_count = tab.max_row

        for row_idx in range(2, row_count + 1):
            name = tab.cell(row = row_idx, column=2).value or ''
            acronym = tab.cell(row = row_idx, column=3).value or ''
            webpage = tab.cell(row = row_idx, column=4).value or ''
            description = tab.cell(row = row_idx, column=5).value or ''
            if acronym not in current_institutions:
                current_institutions.append(acronym)
                new_institution  = Institution(name=name, acronym=acronym, webpage=webpage, description=description)
                new_institution.save()

    schools = School.objects.all()
    current_schools = []
    for school in schools:
        current_schools.append(school.name)

    if 'Schools' in sheet_names:
        tab = xlsx['Schools']
        row_count = tab.max_row

        for row_idx in range(2, row_count + 1):
            institution_str = tab.cell(row = row_idx, column=2).value or ''
            if institution_str not in current_institutions:
                print(f'There is not aronym: {institution_str} in Institutions')
            institution = Institution.objects.get(acronym=institution_str)
            name = tab.cell(row = row_idx, column=3).value or ''
            acronym = tab.cell(row = row_idx, column=4).value or ''
            webpage = tab.cell(row = row_idx, column=5).value or ''
            description = tab.cell(row = row_idx, column=6).value or ''
            if name not in current_schools:
                current_schools.append(name)
                new_school  = School(name=name, acronym=acronym, institution=institution, webpage=webpage, description=description)
                new_school.save()



    departments = Department.objects.all()
    current_departments = []
    for department in departments:
        current_departments.append(department.name)

    if 'Departments' in sheet_names:
        tab = xlsx['Departments']
        row_count = tab.max_row

        for row_idx in range(2, row_count + 1):
            school_str = tab.cell(row = row_idx, column=2).value or ''
            school = School.objects.get(name=school_str)
            name = tab.cell(row = row_idx, column=3).value or ''
            acronym = tab.cell(row = row_idx, column=4).value or ''
            webpage = tab.cell(row = row_idx, column=5).value or ''
            description = tab.cell(row = row_idx, column=6).value or ''

            if name not in current_departments:
                current_departments.append(name)
                new_department  = Department(name=name, acronym=acronym, school=school, webpage=webpage, description=description)
                new_department.save()

    positions = Position.objects.all()
    current_positions = []
    for position in positions:
        current_positions.append(position.name)

    if 'Positions' in sheet_names:
        tab = xlsx['Positions']
        row_count = tab.max_row

        for row_idx in range(2, row_count + 1):
            position = tab.cell(row = row_idx, column=2).value or ''

            if position not in current_positions:
                current_positions.append(position)
                new_position  = Position(name=position)
                new_position.save()


    experts = Expert.objects.all()
    current_experts = []
    for expert in experts:
        current_experts.append(expert.user.email)

    if 'Experts' in sheet_names:
        tab = xlsx["Experts"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            read_idx = tab.cell(row = row_idx, column=1).value
            first_name = tab.cell(row = row_idx, column=2).value or ''
            last_name = tab.cell(row = row_idx, column=3).value or ''
            orcid = tab.cell(row = row_idx, column=4).value or ''

            position_str = tab.cell(row = row_idx, column=5).value
            position = Position.objects.get(name=position_str)

            department_str = tab.cell(row = row_idx, column=6).value
            department = Department.objects.get(name=department_str)

            school_str = tab.cell(row = row_idx, column=7).value
            school = School.objects.get(name=school_str)

            email = tab.cell(row = row_idx, column=8).value
            username = email.split("@")[0]
            phone = tab.cell(row = row_idx, column=9).value
            education = tab.cell(row = row_idx, column=10).value
            expertise = tab.cell(row = row_idx, column=11).value
            biography = tab.cell(row = row_idx, column=12).value
            selected_papers = tab.cell(row = row_idx, column=13).value
            professional_activities = tab.cell(row = row_idx, column=14).value

            if not email in emails and len(email) > 0:
                emails.append(email)
                print(f'USER: {first_name} {last_name} {email}')
                new_user= User(first_name=first_name, last_name=last_name, email=email, username=username)
                new_user.save()
                print(f'(Expert: orcid: {orcid}, position: {position}, phone: {phone}, education: {education}, expertise: {expertise}, biography: {biography}, selected_papers: {selected_papers}, professional_activities: {professional_activities})')
                new_expert = Expert(user=new_user, orcid=orcid, position=position, department=department, phone=phone, expertise=expertise, biography=biography, selected_products=selected_papers, professional_activities=professional_activities)
                new_expert.save()

            else:
                print(read_idx, username, first_name, last_name)
